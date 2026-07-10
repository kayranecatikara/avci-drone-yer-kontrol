# -*- coding: utf-8 -*-
"""TUNE RAPORU (Excel) — ucus logundan gorsel-faz metrikleri + tune degerleri."""

import csv
import glob
import json
import math
import os
import shutil
import time


# Yardimcilar
def _f(x):
    """CSV hucresi -> float | None."""
    if x is None or x == "":
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def en_yeni_log(veri_dir):
    """veri/ucus_log_*.csv icinden en yenisinin yolu (yoksa None)."""
    ler = sorted(glob.glob(os.path.join(veri_dir, "ucus_log_*.csv")),
                 key=os.path.getmtime)
    return ler[-1] if ler else None


def _oku(log_path):
    """Log'u satir listesi olarak oku (DictReader; her deger string)."""
    with open(log_path, "r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


# UCUS KLASORU: bir ucusun tum verileri veri/tune_parametreler/ucus_N/ altinda.
def ucus_klasoru(base_dir, log_path):
    """Ucus loguna karsilik gelen (gerekirse YENI) klasoru dondur."""
    os.makedirs(base_dir, exist_ok=True)
    kayit_yolu = os.path.join(base_dir, "kayit.json")
    try:
        with open(kayit_yolu, "r", encoding="utf-8") as f:
            kayit = json.load(f)
    except Exception:
        kayit = {}
    anahtar = os.path.basename(log_path) if log_path else "logsuz"
    ad = kayit.get(anahtar)
    if ad is None:
        n = 1 + max([int(a.rsplit("_", 1)[1]) for a in kayit.values()
                     if a.rsplit("_", 1)[-1].isdigit()] + [0])
        ad = "ucus_%d" % n
        kayit[anahtar] = ad
        with open(kayit_yolu, "w", encoding="utf-8") as f:
            json.dump(kayit, f, indent=1, ensure_ascii=False)
    klasor = os.path.join(base_dir, ad)
    os.makedirs(klasor, exist_ok=True)
    return klasor


def dosyayi_klasore_al(kaynak, klasor):
    """Log kopyasini ucus klasorune tazele."""
    if not kaynak or not os.path.isfile(kaynak):
        return None
    hedef = os.path.join(klasor, os.path.basename(kaynak))
    try:
        shutil.copy2(kaynak, hedef)
        return hedef
    except Exception:
        return None


def _oku_tune_log(path):
    """1 Hz tune logu -> [(t_wall, {param: deger}), ...] (zaman sirali)."""
    if not path or not os.path.isfile(path):
        return []
    out = []
    with open(path, "r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            t = _f(r.get("t_wall"))
            if t is None:
                continue
            out.append((t, {k: _f(v) for k, v in r.items() if k != "t_wall"}))
    out.sort(key=lambda x: x[0])
    return out


# Metrik hesabi
def metrik_hesapla(satirlar, kilit_gerek_s):
    """Ucus logundan gorsel-faz metrikleri -> [(baslik, deger, aciklama), ...]."""
    M = []

    def ekle(b, d, a=""):
        M.append((b, d, a))

    if not satirlar:
        ekle("HATA", "log bos", "ucus logu satir icermiyor")
        return M

    t0 = _f(satirlar[0].get("t_perf")) or 0.0
    t_son = _f(satirlar[-1].get("t_perf")) or t0
    ekle("Log satiri", len(satirlar), "toplam tik sayisi")
    ekle("Ucus suresi (s)", round(t_son - t0, 1), "ilk->son log satiri")

    vis = [r for r in satirlar if r.get("phase") == "VISUAL"]
    ekle("Gorsel faz tik", len(vis), "phase=VISUAL satir sayisi")
    if not vis:
        ekle("UYARI", "gorsel faz yok",
             "bu ucusta GORSEL_GUDUM hic calismadi; metrikler hesaplanamadi")
        return M
    tv0 = _f(vis[0].get("t_perf")) or t0
    tv1 = _f(vis[-1].get("t_perf")) or tv0
    ekle("Gorsel faz suresi (s)", round(tv1 - tv0, 1), "ilk->son VISUAL satiri")

    # --- 1) HEDEFI NE KADAR HIZLI BULMUS ------------------------------------
    ilk_tespit = next((r for r in satirlar if _f(r.get("vis_gordu")) == 1.0), None)
    if ilk_tespit is not None:
        ti = _f(ilk_tespit.get("t_perf")) or t0
        ekle("Ilk tespit suresi (s)", round(ti - t0, 1),
             "gorev basindan ilk YOLO tespitine")
        gm = _f(ilk_tespit.get("gercek_mesafe"))
        if gm is not None:
            ekle("Ilk tespit mesafesi (m)", round(gm / 100.0, 1),
                 "ilk tespit anindaki gercek 3B mesafe")
    else:
        ekle("Ilk tespit", "YOK", "ucus boyunca hic tespit olmadi")

    # --- 2) TESPIT SUREKLILIGI / KAYIPLAR (gorsel fazda) ---------------------
    gordu = [(_f(r.get("t_perf")) or 0.0, _f(r.get("vis_gordu")) == 1.0) for r in vis]
    n_gordu = sum(1 for _, g in gordu if g)
    ekle("Tespit orani (%)", round(100.0 * n_gordu / len(gordu), 1),
         "gorsel faz tiklerinde hedef gorulme yuzdesi")
    kayip, en_uzun, blok_t0, onceki = 0, 0.0, None, False
    for t, g in gordu:
        if g and not onceki:
            blok_t0 = t
        if (not g) and onceki:
            kayip += 1
            en_uzun = max(en_uzun, t - (blok_t0 or t))
        onceki = g
    if onceki and blok_t0 is not None:
        en_uzun = max(en_uzun, gordu[-1][0] - blok_t0)
    ekle("Kayip sayisi", kayip, "gordu->kaybetti gecis adedi (az = tutarli takip)")
    ekle("En uzun kesintisiz takip (s)", round(en_uzun, 1),
         "araliksiz tespit edilen en uzun sure")
    confs = [c for c in (_f(r.get("vis_conf")) for r in vis) if c is not None and c > 0]
    if confs:
        ekle("Ortalama tespit guveni", round(sum(confs) / len(confs), 3),
             "vis_conf ortalamasi (gorsel faz)")

    # --- 3) KILITLENME ISTERI (sartname 6.1.2/6.1.4 sayaci) ------------------
    kw = [k for k in (_f(r.get("kilit_win_s")) for r in vis) if k is not None]
    kmax = max(kw) if kw else 0.0
    ekle("Kilit penceresi max (s)", round(kmax, 2),
         "10 sn penceredeki kumulatif kilit suresinin zirvesi")
    ekle("Kilit isteri (>=%.0f s)" % kilit_gerek_s,
         "SAGLANDI" if kmax >= kilit_gerek_s else "SAGLANMADI",
         "sartname kilitlenme kosulu")
    if kmax >= kilit_gerek_s:
        tk = next((_f(r.get("t_perf")) for r in vis
                   if (_f(r.get("kilit_win_s")) or 0.0) >= kilit_gerek_s), None)
        if tk is not None:
            ekle("Kilide ulasma suresi (s)", round(tk - t0, 1),
                 "gorev basindan kilit isterinin saglanmasina")
    kilit_toplam = sum(1 for r in vis if (_f(r.get("kilit_win_s")) or 0.0) > 0)
    ekle("Kilit sayaci aktif tik (%)",
         round(100.0 * kilit_toplam / len(vis), 1),
         "kilit penceresinde sure biriktiren tik orani")

    # --- 4) MERKEZLEME KALITESI (IBVS cizgisi) --------------------------------
    exs = [e for e in (_f(r.get("vis_ex")) for r in vis) if e is not None]
    eys = [e for e in (_f(r.get("vis_ey")) for r in vis) if e is not None]
    rs = [x for x in (_f(r.get("ibvs_r")) for r in vis) if x is not None]

    def _ort(v):
        return sum(v) / len(v) if v else None

    def _std(v):
        if len(v) < 2:
            return None
        m = _ort(v)
        return math.sqrt(sum((x - m) ** 2 for x in v) / len(v))

    if exs:
        ekle("Yatay sapma ort |ex|", round(_ort([abs(x) for x in exs]), 3),
             "0=tam nisan noktasinda, 1=kadraj kenari")
    if eys:
        ekle("Dikey sapma ort |ey|", round(_ort([abs(x) for x in eys]), 3), "")
    if rs:
        ekle("IBVS hata buyuklugu ort r", round(_ort(rs), 3),
             "nisan->bbox cizgi buyuklugu (kucuk = merkezde)")
        ekle("Merkezde kalma (r<0.15) (%)",
             round(100.0 * sum(1 for x in rs if x < 0.15) / len(rs), 1),
             "hedefin nisan noktasina yakin tutuldugu tik orani")

    # --- 5) HAREKET TUTARLILIGI (komut puruzlulugu / salinim) -----------------
    yc = [y for y in (_f(r.get("yaw_cmd")) for r in vis) if y is not None]
    if len(yc) >= 2:
        d_yaw = [abs(yc[i] - yc[i - 1]) for i in range(1, len(yc))]
        ekle("Yaw puruzlulugu ort |dYaw|", round(_ort(d_yaw), 4),
             "tikler arasi komut degisimi (kucuk = akici)")
        s = _std(yc)
        if s is not None:
            ekle("Yaw komut std", round(s, 3), "buyuk = salinimli yaw")
    if len(exs) >= 3 and (tv1 - tv0) > 1.0:
        flip = sum(1 for i in range(1, len(exs))
                   if exs[i] * exs[i - 1] < 0 and abs(exs[i]) > 0.05)
        ekle("Yatay salinim (isaret degisimi/dk)",
             round(60.0 * flip / (tv1 - tv0), 1),
             "hedefin merkez etrafinda sag-sol gidip gelmesi (kucuk = tutarli)")

    # --- 6) YAKLASMA / VURUS ---------------------------------------------------
    mes = [(t, m / 100.0) for t, m in
           (((_f(r.get("t_perf")) or 0.0), _f(r.get("gercek_mesafe"))) for r in vis)
           if m is not None]
    if mes:
        m0, msn = mes[0][1], mes[-1][1]
        mmin = min(m for _, m in mes)
        ekle("Gorsel faz baslangic mesafesi (m)", round(m0, 1), "gercek 3B (truth)")
        ekle("Minimum mesafe (m)", round(mmin, 2),
             "hedefe en cok yaklasilan an (vurus kaniti)")
        ekle("Son mesafe (m)", round(msn, 1), "log sonundaki mesafe")
        dt = mes[-1][0] - mes[0][0]
        if dt > 1.0:
            ekle("Ortalama kapanma hizi (m/s)", round((m0 - msn) / dt, 2),
                 "pozitif = hedefe yaklasiyor")

    # --- 7) ONGORULU YAW LEAD (pose) -------------------------------------------
    rok = [x for x in (_f(r.get("ibvs_roll_ok")) for r in vis) if x is not None]
    if rok:
        ekle("Ongoru aktif tik (%)", round(100.0 * sum(1 for x in rok if x == 1.0)
                                           / len(rok), 1),
             "pose kanat-ucu kapilarinin acik oldugu oran")
    leads = [abs(x) for x in (_f(r.get("ibvs_lead")) for r in vis)
             if x is not None and x != 0.0]
    if leads:
        ekle("Ongoru katkisi ort |lead|", round(_ort(leads), 4),
             "yaw'a eklenen ileri-besleme buyuklugu")

    return M


# SEGMENT KIYASI: tune degisim noktalari ucusu segmentlere boler (ucus-ici A/B).
_SEG_BIRLESTIR_S = 3.0     # bu kadar yakin degisimler tek segment sayilir


def _ort(v):
    return sum(v) / len(v) if v else None


def _degisimler(tune_rows):
    """Ardisik tune satirlarini karsilastir -> [(t, 'PARAM: eski->yeni; ...'), ...]"""
    out = []
    for (t0, v0), (t1, v1) in zip(tune_rows, tune_rows[1:]):
        d = []
        for k in sorted(v1):
            a, b = v0.get(k), v1.get(k)
            if a is not None and b is not None and abs(a - b) > 1e-9:
                d.append("%s: %g->%g" % (k, a, b))
        if d:
            out.append((t1, "; ".join(d)))
    # slider suruklemesi = ayni parametrede pespese degisim -> tek noktaya birlestir
    birlesik = []
    for t, d in out:
        if birlesik and t - birlesik[-1][0] <= _SEG_BIRLESTIR_S:
            birlesik[-1] = (t, birlesik[-1][1] + "; " + d)
        else:
            birlesik.append((t, d))
    return birlesik


def _segment_metrik(vis):
    """Bir zaman diliminin VISUAL satirlarindan kompakt metrik sozlugu."""
    m = {"gorsel_tik": len(vis)}
    if not vis:
        return m
    gordu = [_f(r.get("vis_gordu")) == 1.0 for r in vis]
    m["tespit_%"] = round(100.0 * sum(gordu) / len(gordu), 1)
    kayip = sum(1 for a, b in zip(gordu, gordu[1:]) if a and not b)
    m["kayip"] = kayip
    confs = [c for c in (_f(r.get("vis_conf")) for r in vis) if c is not None and c > 0]
    m["conf_ort"] = round(_ort(confs), 3) if confs else None
    rs = [x for x in (_f(r.get("ibvs_r")) for r in vis) if x is not None]
    if rs:
        m["r_ort"] = round(_ort(rs), 3)
        m["merkez_%"] = round(100.0 * sum(1 for x in rs if x < 0.15) / len(rs), 1)
    kw = [k for k in (_f(r.get("kilit_win_s")) for r in vis) if k is not None]
    m["kilit_max_s"] = round(max(kw), 2) if kw else 0.0
    yc = [y for y in (_f(r.get("yaw_cmd")) for r in vis) if y is not None]
    if len(yc) >= 2:
        m["yaw_przlk"] = round(_ort([abs(b - a) for a, b in zip(yc, yc[1:])]), 4)
    mes = [(_f(r.get("t_wall")) or 0.0, _f(r.get("gercek_mesafe")))
           for r in vis if _f(r.get("gercek_mesafe")) is not None]
    if mes:
        m["mesafe0_m"] = round(mes[0][1] / 100.0, 1)
        m["mesafe1_m"] = round(mes[-1][1] / 100.0, 1)
        m["min_mesafe_m"] = round(min(x for _, x in mes) / 100.0, 2)
        dt = mes[-1][0] - mes[0][0]
        if dt > 1.0:
            m["kapanma_mps"] = round((mes[0][1] - mes[-1][1]) / 100.0 / dt, 2)
    return m


_SEG_KOLON = ["gorsel_tik", "tespit_%", "kayip", "conf_ort", "r_ort", "merkez_%",
              "kilit_max_s", "yaw_przlk", "mesafe0_m", "mesafe1_m", "min_mesafe_m",
              "kapanma_mps"]
_SEG_ACIKLAMA = ("tespit_% yuksek + kayip az + r_ort kucuk + merkez_% yuksek + "
                 "yaw_przlk kucuk + kapanma_mps buyuk = IYI segment")


def segment_tablosu(satirlar, tune_rows):
    """Ucusu tune degisimlerine gore segmentlere bol -> (basliklar, satirlar)."""
    if not satirlar:
        return [], []
    tw = [r for r in satirlar if _f(r.get("t_wall")) is not None]
    if not tw:
        return [], []
    t0 = _f(tw[0]["t_wall"])
    t_son = _f(tw[-1]["t_wall"])
    degisim = [(t, d) for t, d in _degisimler(tune_rows) if t0 < t < t_son]
    sinir = [t0] + [t for t, _ in degisim] + [t_son + 0.001]
    params = sorted(tune_rows[0][1].keys()) if tune_rows else []

    basliklar = (["Seg", "Baslangic (s)", "Sure (s)", "Degisiklik"]
                 + _SEG_KOLON + params)
    rows = []
    for i in range(len(sinir) - 1):
        a, b = sinir[i], sinir[i + 1]
        vis = [r for r in tw if r.get("phase") == "VISUAL"
               and a <= _f(r["t_wall"]) < b]
        m = _segment_metrik(vis)
        # segment icindeki SON tune satiri = bu dilimde gecerli parametre seti
        pv = {}
        for t, vals in tune_rows:
            if t < b:
                pv = vals
            else:
                break
        rows.append([i + 1, round(a - t0, 1), round(b - a, 1),
                     "(baslangic)" if i == 0 else degisim[i - 1][1]]
                    + [m.get(k, "") for k in _SEG_KOLON]
                    + [pv.get(k, "") for k in params])
    return basliklar, rows


def saniye_tablosu(satirlar, tune_rows):
    """Saniye bazli gorsel metrik + o saniyedeki tune degerleri -> (basliklar, satirlar)."""
    tw = [r for r in satirlar if _f(r.get("t_wall")) is not None]
    if not tw:
        return [], []
    t0 = _f(tw[0]["t_wall"])
    params = sorted(tune_rows[0][1].keys()) if tune_rows else []
    basliklar = (["t (s)", "faz", "tespit_%", "conf_ort", "|ex|_ort", "|ey|_ort",
                  "r_ort", "kilit_win_s", "mesafe_m", "yaw_cmd_ort"] + params)

    # saniye kovalari (int t_wall)
    kova = {}
    for r in tw:
        kova.setdefault(int(_f(r["t_wall"])), []).append(r)

    rows, ti = [], 0
    for sn in sorted(kova):
        grup = kova[sn]
        vis = [r for r in grup if r.get("phase") == "VISUAL"]
        faz = max(set(r.get("phase") for r in grup),
                  key=lambda p: sum(1 for r in grup if r.get("phase") == p))
        gordu = [_f(r.get("vis_gordu")) == 1.0 for r in vis]
        confs = [c for c in (_f(r.get("vis_conf")) for r in vis) if c is not None and c > 0]
        exs = [abs(x) for x in (_f(r.get("vis_ex")) for r in vis) if x is not None]
        eys = [abs(x) for x in (_f(r.get("vis_ey")) for r in vis) if x is not None]
        rs = [x for x in (_f(r.get("ibvs_r")) for r in vis) if x is not None]
        kw = [k for k in (_f(r.get("kilit_win_s")) for r in vis) if k is not None]
        mes = [m for m in (_f(r.get("gercek_mesafe")) for r in grup) if m is not None]
        yc = [y for y in (_f(r.get("yaw_cmd")) for r in vis) if y is not None]
        # bu saniyede gecerli tune seti: t_wall <= saniye sonu olan SON satir
        while ti + 1 < len(tune_rows) and tune_rows[ti + 1][0] <= sn + 1:
            ti += 1
        pv = tune_rows[ti][1] if tune_rows and tune_rows[ti][0] <= sn + 1 else {}

        def _r(v, n=3):
            return round(v, n) if v is not None else ""
        rows.append([round(sn - t0, 0), faz,
                     _r(100.0 * sum(gordu) / len(gordu), 1) if gordu else "",
                     _r(_ort(confs)), _r(_ort(exs)), _r(_ort(eys)), _r(_ort(rs)),
                     _r(max(kw), 2) if kw else "",
                     _r(_ort(mes) / 100.0, 1) if mes else "",
                     _r(_ort(yc))]
                    + [pv.get(k, "") for k in params])
    return basliklar, rows


# ----------------------------------------------------------------
#  Excel yazimi
# ----------------------------------------------------------------
def rapor_uret(tune_vals, sabit_vals, log_path, cikti_dir, kilit_gerek_s=5.0,
               tune_log_path=None):
    """Tune degerleri + ucus metriklerini xlsx'e yaz -> (xlsx_yolu, ozet_dict)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    satirlar = _oku(log_path) if (log_path and os.path.isfile(log_path)) else []
    metrikler = metrik_hesapla(satirlar, kilit_gerek_s) if satirlar else \
        [("UYARI", "ucus logu yok", "Cfg.LOG_ENABLE=True ile ucus yapip tekrar dene")]

    wb = Workbook()
    kalin = Font(bold=True)

    def sayfa_yaz(ws, basliklar, satirlar_):
        ws.append(basliklar)
        for c in ws[1]:
            c.font = kalin
        for s in satirlar_:
            ws.append(list(s))
        for i, b in enumerate(basliklar, start=1):
            gen = max([len(str(b))] + [len(str(r[i - 1])) for r in satirlar_ if len(r) >= i])
            ws.column_dimensions[get_column_letter(i)].width = min(gen + 3, 80)

    # Sayfa 1: OZET (ne / ne zaman / hangi log)
    ws = wb.active
    ws.title = "Ozet"
    sayfa_yaz(ws, ["Alan", "Deger"], [
        ("Rapor zamani", time.strftime("%Y-%m-%d %H:%M:%S")),
        ("Ucus logu", os.path.basename(log_path) if log_path else "YOK"),
        ("Aciklama", "Tune panelindeki canli degerler + bu ucusun gorsel-faz "
                     "performans metrikleri. Ayni degerlerle kosulari kiyasla."),
    ])

    # Sayfa 2: TUNE DEGERLERI (slider seti + sabitler)
    ws = wb.create_sheet("Tune Degerleri")
    rows = [(k, v, "slider (canli)") for k, v in sorted(tune_vals.items())]
    rows += [(k, v, "sabit (Cfg)") for k, v in sorted(sabit_vals.items())]
    sayfa_yaz(ws, ["Parametre", "Deger", "Kaynak"], rows)

    # Sayfa 3: PERFORMANS
    ws = wb.create_sheet("Performans")
    sayfa_yaz(ws, ["Metrik", "Deger", "Aciklama"], metrikler)

    # Sayfa 4-5: SEGMENT KIYAS + SANIYE DETAY (tune logu varsa)
    tune_rows = _oku_tune_log(tune_log_path)
    seg_sayisi = 0
    if satirlar and tune_rows:
        bas, rows = segment_tablosu(satirlar, tune_rows)
        if rows:
            seg_sayisi = len(rows)
            ws = wb.create_sheet("Segment Kiyas")
            ws.append([_SEG_ACIKLAMA])
            ws["A1"].font = Font(italic=True)
            ws.append([])
            ws.append(bas)
            for c in ws[3]:
                c.font = kalin
            for r in rows:
                ws.append(r)
            for i in range(1, len(bas) + 1):
                ws.column_dimensions[get_column_letter(i)].width = \
                    max(10, min(len(str(bas[i - 1])) + 2, 60))
            ws.column_dimensions["D"].width = 44        # Degisiklik kolonu genis
        bas, rows = saniye_tablosu(satirlar, tune_rows)
        if rows:
            ws = wb.create_sheet("Saniye Detay")
            sayfa_yaz(ws, bas, rows)

    os.makedirs(cikti_dir, exist_ok=True)
    yol = os.path.join(cikti_dir, time.strftime("tune_rapor_%Y%m%d_%H%M%S.xlsx"))
    wb.save(yol)

    ozet = {b: d for b, d, _ in metrikler
            if b in ("Ilk tespit suresi (s)", "Tespit orani (%)", "Kayip sayisi",
                     "En uzun kesintisiz takip (s)", "Kilit penceresi max (s)",
                     "IBVS hata buyuklugu ort r", "Minimum mesafe (m)")}
    if seg_sayisi > 1:
        ozet["Tune segmenti"] = "%d (Segment Kiyas sayfasi)" % seg_sayisi
    return yol, ozet
