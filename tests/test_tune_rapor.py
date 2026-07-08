# -*- coding: utf-8 -*-
"""
TUNE RAPORU testleri (web/tune_rapor.py — "Degerleri Yazdir" Excel raporu).

Sentetik bir ucus logu uretilir (0-5 sn APPROACH tespitsiz; 5-35 sn VISUAL,
12-13.5 sn arasi bir kayip blogu, kilit penceresi 14. sn'den itibaren dolar)
ve metriklerin bu kurgudan dogru cikarildigi + xlsx'in 3 sayfayla yazildigi
dogrulanir. Calistir:  python tests/test_tune_rapor.py
"""

import csv
import math
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from web.tune_rapor import (rapor_uret, en_yeni_log,         # noqa: E402
                            ucus_klasoru, dosyayi_klasore_al)
from guidance.ana_kontrol import _LOG_COLS                   # noqa: E402

_SONUC = []


def _test(ad, kosul, detay=""):
    _SONUC.append((ad, bool(kosul)))
    print(("OK  " if kosul else "FAIL") + " " + ad + (("  [" + str(detay) + "]") if not kosul else ""))


_T_WALL0 = 1000000.0        # sentetik duvar-saati taban (t_wall = taban + t_perf)


def _sentetik_log(yol):
    rows, t = [], 0.0
    while t < 5.0:                                           # tespitsiz yaklasma
        r = {k: "" for k in _LOG_COLS}
        r.update({"t_perf": t, "t_wall": _T_WALL0 + t, "phase": "APPROACH",
                  "gercek_mesafe": 12000 - t * 400, "yaw_cmd": 0.1})
        rows.append(r); t += 0.1
    while t < 35.0:                                          # gorsel faz
        gordu = 0 if 12.0 <= t < 13.5 else 1                 # tek kayip blogu
        # 20. sn'de "tune iyilesmesi": sapma ve puruzluluk kuculur (segment kiyasinda gorunmeli)
        iyi = 0.4 if t >= 20.0 else 1.0
        ex = 0.3 * iyi * math.sin(t * 0.8) * gordu
        ey = 0.1 * iyi * math.cos(t * 0.5)
        r = {k: "" for k in _LOG_COLS}
        r.update({"t_perf": t, "t_wall": _T_WALL0 + t, "phase": "VISUAL",
                  "vis_gordu": gordu,
                  "vis_conf": 0.62 if gordu else "", "vis_ex": ex, "vis_ey": ey,
                  "ibvs_r": math.hypot(ex, ey),
                  "kilit_win_s": max(0.0, min(6.0, t - 14.0)),
                  "gercek_mesafe": max(150, 10000 - (t - 5) * 320),
                  "yaw_cmd": 0.4 * ex, "ibvs_roll_ok": gordu,
                  "ibvs_lead": 0.05 * gordu})
        rows.append(r); t += 0.1
    with open(yol, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=_LOG_COLS)
        w.writeheader(); w.writerows(rows)


def _sentetik_tune_log(yol):
    # 1 Hz: IBVS_K_YAW 20. sn'de 0.8 -> 1.2 (ucus ortasi degisim -> 2 segment)
    with open(yol, "w", encoding="utf-8") as f:
        f.write("t_wall,IBVS_ILERI,IBVS_K_YAW\n")
        for sn in range(0, 36):
            k_yaw = 1.2 if sn >= 20 else 0.8
            f.write("%.3f,0.35,%g\n" % (_T_WALL0 + sn, k_yaw))


def main():
    tmp = tempfile.mkdtemp()
    log = os.path.join(tmp, "ucus_log_test.csv")
    tlog = os.path.join(tmp, "tune_log_test.csv")
    _sentetik_log(log)
    _sentetik_tune_log(tlog)

    yol, ozet = rapor_uret({"IBVS_K_YAW": 0.8, "IBVS_ILERI": 0.35},
                           {"IBVS_SIGN_YAW": 1.0, "VIS_WIN_NEED_S": 5.0},
                           log, tmp, kilit_gerek_s=5.0, tune_log_path=tlog)
    _test("xlsx_dosyasi_yazildi", os.path.isfile(yol), yol)
    _test("ozet_dolu", len(ozet) >= 5, ozet)

    from openpyxl import load_workbook
    wb = load_workbook(yol)
    _test("bes_sayfa", wb.sheetnames == ["Ozet", "Tune Degerleri", "Performans",
                                         "Segment Kiyas", "Saniye Detay"],
          wb.sheetnames)

    # --- SEGMENT KIYAS: 20. sn'deki K_YAW degisimi ucusu 2 segmente bolmeli,
    #     ikinci segmentte r_ort belirgin kucuk (sentetik 'iyilesme') olmali. ---
    seg = list(wb["Segment Kiyas"].iter_rows(min_row=3, values_only=True))
    bas, seg_rows = list(seg[0]), seg[1:]
    _test("iki_segment", len(seg_rows) == 2, len(seg_rows))
    i_r = bas.index("r_ort"); i_d = bas.index("Degisiklik"); i_k = bas.index("IBVS_K_YAW")
    _test("degisiklik_metni", "IBVS_K_YAW: 0.8->1.2" in str(seg_rows[1][i_d]),
          seg_rows[1][i_d])
    _test("segment_param_kolonu", seg_rows[0][i_k] == 0.8 and seg_rows[1][i_k] == 1.2,
          (seg_rows[0][i_k], seg_rows[1][i_k]))
    _test("iyilesme_gorunur", seg_rows[1][i_r] < seg_rows[0][i_r] * 0.6,
          (seg_rows[0][i_r], seg_rows[1][i_r]))

    # --- SANIYE DETAY: her ucus saniyesi bir satir + o saniyenin tune degeri ---
    sd = list(wb["Saniye Detay"].iter_rows(min_row=1, values_only=True))
    sbas, sd_rows = list(sd[0]), sd[1:]
    _test("saniye_satir_sayisi", 33 <= len(sd_rows) <= 36, len(sd_rows))
    j_t = sbas.index("t (s)"); j_k = sbas.index("IBVS_K_YAW")
    onceler = [r[j_k] for r in sd_rows if r[j_t] is not None and r[j_t] < 19]
    sonralar = [r[j_k] for r in sd_rows if r[j_t] is not None and r[j_t] >= 21]
    _test("saniye_tune_hizalama", all(v == 0.8 for v in onceler)
          and all(v == 1.2 for v in sonralar),
          (set(onceler), set(sonralar)))
    perf = {r[0]: r[1] for r in wb["Performans"].iter_rows(min_row=2, values_only=True)}
    tune = {r[0]: r[1] for r in wb["Tune Degerleri"].iter_rows(min_row=2, values_only=True)}

    _test("tune_degerleri_sayfada", tune.get("IBVS_K_YAW") == 0.8
          and tune.get("IBVS_SIGN_YAW") == 1.0, tune)
    _test("ilk_tespit_5sn", abs(perf["Ilk tespit suresi (s)"] - 5.0) < 0.3,
          perf["Ilk tespit suresi (s)"])
    _test("kayip_sayisi_1", perf["Kayip sayisi"] == 1, perf["Kayip sayisi"])
    _test("kilit_isteri_saglandi", perf["Kilit isteri (>=5 s)"] == "SAGLANDI")
    _test("tespit_orani_yuzde", 85 < perf["Tespit orani (%)"] < 100,
          perf["Tespit orani (%)"])
    _test("min_mesafe_metre", 3.5 < perf["Minimum mesafe (m)"] < 4.5,
          perf["Minimum mesafe (m)"])

    # log yokken cokmemeli (uyari satiriyla yine xlsx uretir)
    yol2, _ = rapor_uret({}, {}, None, tmp)
    _test("log_yokken_cokmez", os.path.isfile(yol2))
    # en_yeni_log bos dizinde None doner
    _test("en_yeni_log_bos_dizin", en_yeni_log(os.path.join(tmp, "bos")) is None)

    # --- UCUS KLASORU: ayni log -> ayni klasor; yeni log -> siradaki numara ---
    base = os.path.join(tmp, "tune_parametreler")
    k1 = ucus_klasoru(base, log)
    k1b = ucus_klasoru(base, log)
    _test("ayni_log_ayni_klasor", k1 == k1b and os.path.basename(k1) == "ucus_1",
          (k1, k1b))
    log2 = os.path.join(tmp, "ucus_log_ikinci.csv")
    _sentetik_log(log2)
    k2 = ucus_klasoru(base, log2)
    _test("yeni_log_yeni_klasor", os.path.basename(k2) == "ucus_2", k2)
    # kopyalama: log + tune logu klasore duser; rapor da klasore yazilir
    dosyayi_klasore_al(log, k1)
    dosyayi_klasore_al(tlog, k1)
    yol3, _ = rapor_uret({"IBVS_K_YAW": 0.8}, {}, log, k1, tune_log_path=tlog)
    icerik = set(os.listdir(k1))
    _test("klasor_icerigi_tam",
          os.path.basename(log) in icerik and os.path.basename(tlog) in icerik
          and os.path.basename(yol3) in icerik, icerik)

    n_ok = sum(1 for _, ok in _SONUC if ok)
    print(f"\n{n_ok}/{len(_SONUC)} test gecti.")
    sys.exit(0 if n_ok == len(_SONUC) else 1)


if __name__ == "__main__":
    main()
